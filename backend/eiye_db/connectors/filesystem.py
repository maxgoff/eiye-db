"""Filesystem connector: CSV, text, PDF, and XLSX files under a configured root."""

import asyncio
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from eiye_db.connectors.base import Connector, ConnectorError

_TEXT_SUFFIXES = {".txt", ".md", ".log", ".json"}
_MAX_FILES = 1000
_SAMPLE_ROWS = 20
_MAX_TEXT_CHARS = 100_000
_MAX_XLSX_COLS = 256
_MAX_CELL_CHARS = 10_000
_MAX_XLSX_INSPECT_BYTES = 5_000_000  # skip xlsx field inference above this during discovery


def infer_type(values: list[str]) -> str:
    non_empty = [v for v in values if v != ""]
    if not non_empty:
        return "string"
    try:
        for v in non_empty:
            int(v)
        return "integer"
    except ValueError:
        pass
    try:
        for v in non_empty:
            float(v)
        return "number"
    except ValueError:
        return "string"


class FilesystemConnector(Connector):
    def _root(self) -> Path:
        root = self.config.get("root")
        if not root:
            raise ConnectorError("filesystem config requires 'root'")
        return Path(root).resolve()

    def _resolve(self, rel_path: str) -> Path:
        root = self._root()
        target = (root / rel_path).resolve()
        if root != target and root not in target.parents:
            raise ConnectorError(f"path escapes datasource root: {rel_path}")
        return target

    async def test_connection(self) -> None:
        root = self._root()
        if not root.is_dir():
            raise ConnectorError(f"root is not a directory: {root}")

    async def discover_schema(self) -> list[dict[str, Any]]:
        root = self._root()
        if not root.is_dir():
            raise ConnectorError(f"root is not a directory: {root}")
        tables = []
        count = 0
        for path in sorted(root.rglob("*")):
            if not path.is_file():
                continue
            count += 1
            if count > _MAX_FILES:
                break
            rel = str(path.relative_to(root))
            suffix = path.suffix.lower()
            if suffix == ".csv":
                tables.append({"name": rel, "fields": self._csv_fields(path)})
            elif suffix == ".xlsx":
                tables.append({"name": rel, "fields": self._xlsx_fields(path)})
            elif suffix == ".pdf" or suffix in _TEXT_SUFFIXES:
                tables.append({"name": rel, "fields": [{"name": "content", "type": "text"}]})
        return tables

    def _csv_fields(self, path: Path) -> list[dict[str, Any]]:
        try:
            with path.open(newline="", errors="replace") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if not header:
                    return []
                samples: list[list[str]] = [[] for _ in header]
                for i, row in enumerate(reader):
                    if i >= _SAMPLE_ROWS:
                        break
                    for col, value in enumerate(row[: len(header)]):
                        samples[col].append(value)
            return [{"name": h, "type": infer_type(samples[i])} for i, h in enumerate(header)]
        except OSError as e:
            raise ConnectorError(f"cannot read {path.name}: {e}") from e

    def _xlsx_fields(self, path: Path) -> list[dict[str, Any]]:
        # Best-effort field inference from the active sheet's header + sample rows.
        # Returns [] rather than failing discovery if the workbook can't be read.
        # Skip very large workbooks so discovery never has to parse a huge file.
        try:
            if path.stat().st_size > _MAX_XLSX_INSPECT_BYTES:
                return []
        except OSError:
            return []
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception:
            return []
        try:
            ws = wb.active
            if ws is None:
                return []
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                return []
            cols = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)][:_MAX_XLSX_COLS]
            samples: list[list[str]] = [[] for _ in cols]
            for i, row in enumerate(rows):
                if i >= _SAMPLE_ROWS:
                    break
                for c, value in enumerate(row[: len(cols)]):
                    samples[c].append("" if value is None else str(value)[:_MAX_CELL_CHARS])
            return [{"name": cols[i], "type": infer_type(samples[i])} for i in range(len(cols))]
        except Exception:
            return []
        finally:
            wb.close()

    def _xlsx_rows(self, path: Path, limit: int) -> list[dict[str, Any]]:
        # Active sheet only; multi-sheet workbooks expose just the active sheet for now.
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            raise ConnectorError(f"cannot read xlsx {path.name}: {e}") from e
        try:
            ws = wb.active
            if ws is None:
                return []
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                return []
            cols = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)][:_MAX_XLSX_COLS]
            out: list[dict[str, Any]] = []
            for row in rows:
                if len(out) >= limit:
                    break
                out.append(
                    {cols[i]: ("" if v is None else str(v)[:_MAX_CELL_CHARS]) for i, v in enumerate(row[: len(cols)])}
                )
            return out
        except ConnectorError:
            raise
        except Exception as e:
            raise ConnectorError(f"cannot read xlsx {path.name}: {e}") from e
        finally:
            wb.close()

    def _pdf_rows(self, path: Path) -> list[dict[str, Any]]:
        try:
            reader = PdfReader(str(path))
        except Exception as e:
            raise ConnectorError(f"cannot read PDF {path.name}: {e}") from e
        if reader.is_encrypted:
            raise ConnectorError(f"encrypted PDF not supported: {path.name}")
        parts: list[str] = []
        total = 0
        try:
            for page in reader.pages:
                text = page.extract_text() or ""
                parts.append(text)
                total += len(text)
                if total >= _MAX_TEXT_CHARS:  # stop before pulling a huge doc fully into memory
                    break
        except Exception as e:
            raise ConnectorError(f"cannot extract text from PDF {path.name}: {e}") from e
        return [{"content": "\n".join(parts)[:_MAX_TEXT_CHARS]}]

    async def query(self, request: dict[str, Any], limit: int) -> list[dict[str, Any]]:
        rel_path = request.get("path")
        if not rel_path:
            raise ConnectorError("filesystem query requires 'path'")
        target = self._resolve(rel_path)
        if not target.is_file():
            raise ConnectorError(f"no such file: {rel_path}")
        suffix = target.suffix.lower()
        try:
            if suffix == ".csv":
                with target.open(newline="", errors="replace") as f:
                    # restkey keeps overflow cells from ragged rows under a
                    # string key instead of None (which breaks JSON serialization).
                    reader = csv.DictReader(f, restkey="_extra")
                    return [row for _, row in zip(range(limit), reader)]
            # PDF/XLSX parsing is blocking CPU work; run it off the event loop so a slow or
            # huge file cannot stall other requests (and the query timeout can still fire).
            if suffix == ".xlsx":
                return await asyncio.to_thread(self._xlsx_rows, target, limit)
            if suffix == ".pdf":
                return await asyncio.to_thread(self._pdf_rows, target)
            return [{"content": target.read_text(errors="replace")[:_MAX_TEXT_CHARS]}]
        except OSError as e:
            raise ConnectorError(f"cannot read {rel_path}: {e}") from e
